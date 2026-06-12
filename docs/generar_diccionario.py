"""Genera docs/diccionario_metricas.xlsx — documentación de todas las métricas.

Uso (desde cualquier carpeta):  python docs/generar_diccionario.py
Regenerar cada vez que se añadan o cambien indicadores, para que el diccionario
no quede desactualizado (la hoja 4 se autogenera desde src/agrupaciones.py).
"""
import os
import sys

_RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _RAIZ)
os.chdir(_RAIZ)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src import agrupaciones as ag

AZUL = "1F4E78"
GRIS = "F2F2F2"
F_TIT = Font(name="Arial", size=14, bold=True, color=AZUL)
F_HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
F_TXT = Font(name="Arial", size=10)
FILL_HDR = PatternFill("solid", start_color=AZUL)
FILL_ALT = PatternFill("solid", start_color=GRIS)
WRAP = Alignment(wrap_text=True, vertical="top")
BORDE = Border(*[Side(style="thin", color="D9D9D9")] * 4)


def hoja_tabla(wb, nombre, titulo, subtitulo, headers, rows, widths):
    ws = wb.create_sheet(nombre)
    ws["A1"] = titulo
    ws["A1"].font = F_TIT
    ws["A2"] = subtitulo
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    hr = 4
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=j, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDE
    for i, row in enumerate(rows):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=hr + 1 + i, column=j, value=val)
            c.font = F_TXT
            c.alignment = WRAP
            c.border = BORDE
            if i % 2 == 1:
                c.fill = FILL_ALT
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{hr + 1}"
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{hr + len(rows)}"
    return ws


wb = Workbook()

# ── Portada ───────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Léeme"
ws.column_dimensions["A"].width = 110
filas = [
    ("Diccionario de métricas — Análisis Financiero Sector Solidario", F_TIT),
    ("", None),
    ("Generado el 2026-06-10 a partir del código del proyecto (src/analytics.py, "
     "src/agrupaciones.py, src/etl.py y views/). Objetivo: revisión de fórmulas.", None),
    ("", None),
    ("CONTENIDO", Font(name="Arial", size=11, bold=True)),
    ("  1. Básicos — 8 indicadores sobre cuentas principales (analytics.agregar_indicadores).", None),
    ("  2. Seis dígitos — 20 indicadores sobre el plan de cuentas a 6 dígitos "
     "(analytics.indicadores_6dig), con fórmulas portadas del proyecto analisis_tasas.", None),
    ("  3. Otras métricas — Panorama, Tasas (ahorro y crédito) y Riesgo.", None),
    ("  4. Agrupaciones — catálogo PUC completo (src/agrupaciones.py): qué cuentas "
     "componen cada concepto usado en las fórmulas de la hoja 2.", None),
    ("", None),
    ("CONVENCIONES", Font(name="Arial", size=11, bold=True)),
    ("  · Todos los indicadores en % salvo que se indique; división por cero devuelve NaN "
     "(la celda queda vacía en la app, nunca error).", None),
    ("  · Códigos de 6 dígitos = cuentas PUC Supersolidaria (ej. 140000 = Cartera de créditos).", None),
    ("  · Corte de datos: 31/03/2026 (CAC tasas y estados: 30/04/2026 según reporte).", None),
    ("  · Los nombres en MAYÚSCULAS_GUIÓN (ej. CARTERA_BRUTA) son agrupaciones: ver hoja 4.", None),
]
for i, (txt, fnt) in enumerate(filas, 1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = fnt or F_TXT
    c.alignment = WRAP

# ── Hoja 1: básicos ───────────────────────────────────────────────────────────
H = ["#", "Código interno", "Nombre en la app", "Fórmula", "Cuentas PUC",
     "Dónde se usa", "Observaciones / advertencias"]
basicos = [
    (1, "ROA", "ROA del periodo", "Excedente / Activo × 100", "350000 / 100000",
     "Explorador, Comparador",
     "Proxy: excedente ACUMULADO del trimestre sobre saldo del activo a la fecha de corte. "
     "No anualizado ni con activo promedio (no hay serie mensual)."),
    (2, "ROE", "ROE del periodo", "Excedente / Patrimonio × 100", "350000 / 300000",
     "Explorador, Comparador", "Mismo proxy trimestral que el ROA."),
    (3, "SOLVENCIA", "Solvencia", "Patrimonio / Activo × 100", "300000 / 100000",
     "Explorador, Comparador, Panorama (agregado del sector)",
     "OJO: es un proxy de capitalización, NO la solvencia regulatoria del Decreto 961/2018 "
     "(patrimonio técnico / APNR), que no se puede calcular con estos reportes."),
    (4, "ENDEUDAMIENTO", "Endeudamiento", "Pasivo / Activo × 100", "200000 / 100000",
     "Explorador, Comparador", "Complemento de la solvencia (suman 100%)."),
    (5, "CARTERA_ACTIVO", "Cartera / Activo", "Cartera / Activo × 100", "140000 / 100000",
     "Explorador, Comparador", "140000 es cartera NETA (ya descuenta provisiones)."),
    (6, "FONDEO_DEPOSITOS", "Fondeo por depósitos", "Depósitos / Activo × 100", "210000 / 100000",
     "Explorador, Comparador",
     "Depósitos BRUTOS (incluyen intereses causados). La versión fina está en la hoja 2 "
     "(FONDEO_DEP_CARTERA usa depósitos netos y cartera bruta)."),
    (7, "MARGEN_EXCEDENTE", "Margen de excedente", "Excedente / Ingresos × 100", "350000 / 400000",
     "Explorador, Comparador", ""),
    (8, "EFICIENCIA", "Carga administrativa", "Gastos de administración / Ingresos × 100",
     "510000 / 400000", "Explorador, Comparador",
     "Versión simple. La eficiencia operativa de la hoja 2 (portada de analisis_tasas) "
     "incluye costos financieros y usa ingresos de cartera como denominador."),
]
hoja_tabla(wb, "1. Básicos",
           "Indicadores básicos (cuentas principales)",
           "src/analytics.py · agregar_indicadores() · fuente: reporte de cuentas principales (1.467 entidades)",
           H, basicos, [4, 22, 22, 38, 18, 26, 60])

# ── Hoja 2: seis dígitos ──────────────────────────────────────────────────────
H6 = ["#", "Código interno", "Nombre en la app", "Fórmula (agrupaciones → hoja 4)",
      "Cuentas PUC principales", "Dónde se usa", "Observaciones / advertencias"]
seis = [
    (1, "CALIDAD_RIESGO", "Calidad por riesgo",
     "CARTERA_EN_RIESGO / CARTERA_INTEGRAL × 100",
     "Subcuentas B–E (capital+intereses+otros) / cartera total con intereses y otros",
     "Explorador, Comparador",
     "En analisis_tasas el denominador era la cartera del archivo de crédito individual; "
     "aquí se usa la cartera integral del balance (equivalente contable). La categoría B "
     "cuenta como 'en riesgo' (estándar Supersolidaria)."),
    (2, "CARTERA_PRODUCTIVA", "Cartera productiva (A+B)",
     "CARTERA_PRODUCTIVA_AB / CARTERA_BRUTA × 100",
     "Subcuentas categorías A y B / capital total de cartera",
     "Explorador, Comparador",
     "La categoría B aparece en productiva Y en riesgo a la vez (así lo define el catálogo)."),
    (3, "COBERTURA_RIESGO", "Cobertura por riesgo",
     "PROVISIONES_TOTAL / CARTERA_EN_RIESGO × 100",
     "(prov. individuales cap+int+otros + general 146800) / cartera B–E",
     "Explorador, Comparador", "Puede superar 100% (sobre-cobertura)."),
    (4, "COBERTURA_GENERAL", "Cobertura general",
     "(PROVISIONES_INDIVIDUALES_CAPITAL + PROVISIONES_GENERALES) / CARTERA_BRUTA × 100",
     "(140800+144500+145800+146500+147100+147900 + 146800) / cartera bruta",
     "Explorador, Comparador", ""),
    (5, "PCT_VIVIENDA", "Mezcla: Vivienda",
     "CARTERA_BRUTA_VIVIENDA / CARTERA_BRUTA × 100", "140400+140500",
     "Explorador (torta de mezcla)", ""),
    (6, "PCT_CONSUMO", "Mezcla: Consumo",
     "(CARTERA_BRUTA_CONSUMO_LIBRANZA + CARTERA_BRUTA_CONSUMO_CAJA) / CARTERA_BRUTA × 100",
     "141100+144100 + 141200+144200", "Explorador (torta de mezcla)", ""),
    (7, "PCT_MICROCREDITO", "Mezcla: Microcrédito",
     "CARTERA_BRUTA_MICROCREDITO / CARTERA_BRUTA × 100", "144800+145400+145500",
     "Explorador (torta de mezcla)", "Incluye la cuenta histórica 144800."),
    (8, "PCT_COMERCIAL", "Mezcla: Comercial",
     "CARTERA_BRUTA_COMERCIAL / CARTERA_BRUTA × 100", "146100+146200",
     "Explorador (torta de mezcla)", ""),
    (9, "PCT_PRODUCTIVO", "Mezcla: Productivo y otros",
     "(CARTERA_BRUTA_PRODUCTIVO + CARTERA_BRUTA_EMPLEADOS) / CARTERA_BRUTA × 100",
     "147600 + 146900", "Explorador (torta de mezcla)",
     "Agrupa modalidad Productivo y la histórica de Empleados."),
    (10, "FONDEO_DEP_CARTERA", "Fondeo depósitos / cartera",
     "DEPOSITOS_NETOS / CARTERA_BRUTA × 100",
     "(210500 + 211000−211095 + 212500−212595) / cartera bruta",
     "Explorador, Comparador",
     "Depósitos SIN intereses causados, sobre cartera bruta (no sobre activo): "
     "responde '¿qué parte de la cartera se fondea con ahorro de asociados?'."),
    (11, "FONDEO_APORTES", "Fondeo aportes / cartera",
     "APORTES_SOCIALES_ASOCIADOS / CARTERA_BRUTA × 100", "(310000 − 311010) / cartera bruta",
     "Explorador, Comparador", "Descuenta aportes en amortización."),
    (12, "DEUDA_ACTIVOS", "(calculado, sin vista aún)",
     "OBLIGACIONES_FINANCIERAS / ACTIVOS × 100", "230000 / 100000",
     "Disponible en analytics; no se muestra en ninguna página todavía",
     "Equivale a obligaciones financieras totales con intereses."),
    (13, "ACTIVOS_IMPRODUCTIVOS", "Activos improductivos",
     "ACTIVO_IMPRODUCTIVO / ACTIVOS × 100",
     "Cartera C–E + intereses/otros de cartera + caja y bancos sin rendimiento (110500+111000) "
     "+ cuentas por cobrar (147300+160000) + activos fijos (170000) + para venta (180000) "
     "+ otros (190000) + inversiones patrimoniales (122600) − provisiones totales",
     "Explorador, Comparador",
     "Misma definición que la pestaña Capital de analisis_tasas."),
    (14, "CAPITAL_INSTITUCIONAL", "Capital institucional",
     "CAPITAL_INSTITUCIONAL / ACTIVOS × 100", "(311010+320000+330000+340000) / 100000",
     "Explorador, Comparador", "Reservas + fondos de destinación específica + superávit + aportes amortizados."),
    (15, "IRREDUCIBLE_SOCIAL", "(calculado, sin vista aún)",
     "CAPITAL_IRREDUCIBLE / CAPITAL_SOCIAL × 100", "311000 / 310000",
     "Disponible en analytics; no se muestra en ninguna página todavía", ""),
    (16, "EFICIENCIA_OPERATIVA", "Eficiencia operativa",
     "(GASTOS_ADMINISTRACION + COSTOS_FINANCIEROS) / (INGRESOS_CARTERA + RECUPERACIONES) × 100",
     "(510000 + 600000) / (415000 + 422500)",
     "Explorador, Comparador",
     "Fórmula del Índice de Eficiencia de analisis_tasas. Puede superar 100% en entidades "
     "cuyo ingreso principal no es la cartera (mediana del sector ≈ 101%): interpretarla "
     "solo en entidades con vocación de crédito. Menor = mejor."),
    (17, "MARGEN_FINANCIERO", "Margen financiero",
     "(INGRESOS_CARTERA − COSTOS_FINANCIEROS) / INGRESOS_CARTERA × 100",
     "(415000 − 600000) / 415000", "Explorador, Comparador",
     "Denominador = ingresos de cartera (415000), NO ingresos totales."),
    (18, "MARGEN_OPERACIONAL", "Margen operacional",
     "(INGRESOS_CARTERA + RECUPERACIONES − COSTOS_FINANCIEROS − GASTOS_ADMINISTRACION) / INGRESOS_CARTERA × 100",
     "(415000 + 422500 − 600000 − 510000) / 415000", "Explorador, Comparador",
     "Puede ser negativo (gastos superan el negocio de intermediación)."),
    (19, "DIVERSIFICACION", "Diversificación de ingresos",
     "(INGRESOS_TOTAL − INGRESOS_CARTERA − RECUPERACIONES) / INGRESOS_TOTAL × 100",
     "(400000 − 415000 − 422500) / 400000", "Explorador, Comparador",
     "Alto = la entidad depende poco de la cartera."),
    (20, "DEPENDENCIA_MARGEN", "(calculado, sin vista aún)",
     "(INGRESOS_CARTERA + RECUPERACIONES) / INGRESOS_TOTAL × 100",
     "(415000 + 422500) / 400000",
     "Disponible en analytics; no se muestra en ninguna página todavía",
     "Complemento de la diversificación (suman 100%)."),
]
hoja_tabla(wb, "2. Seis dígitos",
           "Indicadores sobre el plan de cuentas a 6 dígitos",
           "src/analytics.py · indicadores_6dig() · fórmulas portadas de analisis_tasas/financiero/indicadores.py "
           "· fuente: saldos_6dig.parquet (1.467 entidades × 216 cuentas del catálogo)",
           H6, seis, [4, 24, 26, 50, 42, 28, 60])

# ── Hoja 3: otras métricas ────────────────────────────────────────────────────
HO = ["Página", "Métrica", "Fórmula / origen", "Quién la calcula", "Observaciones"]
otras = [
    ("📈 Panorama", "Activos / Pasivos / Patrimonio / Cartera / Depósitos / Excedentes del sector",
     "Suma simple de la cuenta sobre las entidades filtradas",
     "App (analytics.resumen_sector)", "Reaccionan a los filtros de tipo y departamento."),
    ("📈 Panorama", "Asociados / Empleados / Entidades", "Suma o conteo simple",
     "App (analytics.resumen_sector)", ""),
    ("📈 Panorama", "Solvencia del sector", "Σ Patrimonio / Σ Activo × 100",
     "App (views/panorama.py)", "Agregado ponderado (no promedio de solvencias)."),
    ("📈 Panorama", "Concentración Top 10/50/100",
     "Σ activos de las N mayores / Σ activos total × 100", "App (views/panorama.py)", ""),
    ("📈 Panorama", "Entidades que concentran el 50%",
     "Mínimo N tal que las N mayores sumen ≥ 50% de los activos", "App (views/panorama.py)", ""),
    ("💰 Ahorro y crédito", "Tasa activa / pasiva por entidad (y por modalidad: consumo, "
     "vivienda, comercial, microcrédito, CDAT, permanente, contractual, cuenta de ahorro)",
     "Promedios ponderados reportados en el archivo de tasas de la Delegatura Financiera",
     "SUPERSOLIDARIA (la app solo las muestra)",
     "172 entidades con delegatura financiera. La app no recalcula nada."),
    ("💰 Ahorro y crédito", "MARGEN (margen de intermediación)",
     "TASA_ACTIVA − TASA_PASIVA (puntos porcentuales)", "ETL (src/etl.py construir_tasas)",
     "Resta simple de tasas efectivas; no es margen sobre saldos."),
    ("💰 Ahorro y crédito", "KPIs de cabecera (tasa activa media, pasiva media, margen medio)",
     "Promedio SIMPLE entre entidades filtradas", "App (views/ahorro_credito.py)",
     "Sin ponderar por tamaño: cada entidad pesa igual."),
    ("⚠️ Riesgo", "Indicador de cartera del subsector y su σ",
     "Serie histórica oficial (archivo desviación estándar)",
     "SUPERSOLIDARIA (la app solo la grafica)",
     "Umbrales de alerta = promedio + 1σ y + 2σ; cruzar +2σ = señal de deterioro."),
    ("⚠️ Riesgo", "Factores VaR (medias y desviaciones) y matriz de correlación",
     "Archivo oficial VaR marzo 2026 (TES COP/UVR)",
     "SUPERSOLIDARIA (la app solo los muestra)",
     "Son los INSUMOS del VaR; la app no calcula el VaR de un portafolio."),
    ("(legado)", "src/indicators.py: solvencia, cartera/activo, fondeo, margen, ROA",
     "Mismas fórmulas que la hoja 1, en lógica pura para un balance suelto",
     "Código legado del scaffold", "Solo lo usan los tests; la app usa analytics.py. "
     "Diferencia: división por cero devuelve 0 (no NaN)."),
]
hoja_tabla(wb, "3. Otras métricas",
           "Métricas de Panorama, Ahorro y Crédito, y Riesgo",
           "Incluye las que vienen ya calculadas por la Supersolidaria (la app solo las visualiza)",
           HO, otras, [18, 44, 48, 30, 55])

# ── Hoja 4: agrupaciones (autogenerada desde el código) ──────────────────────
HA = ["Categoría", "Alias", "Nombre", "Tipo", "Definición (cuenta×signo o componentes)", "Descripción"]
rows = []
for a in ag.TODAS:
    if "cuentas" in a:
        tipo = "Cuentas PUC"
        definicion = "  ".join(f"{'+' if s > 0 else '−'}{c}" for c, s in a["cuentas"])
    else:
        tipo = "Compuesta"
        partes = []
        for comp in a["componentes"]:
            if isinstance(comp, str):
                partes.append(f"+{comp}")
            else:
                partes.append(f"{'+' if comp['signo'] > 0 else '−'}{comp['alias']}")
        definicion = "  ".join(partes)
    rows.append((a["categoria"], a["alias"], a["nombre"], tipo, definicion,
                 a.get("descripcion", "")))
hoja_tabla(wb, "4. Agrupaciones",
           "Catálogo de agrupaciones PUC (src/agrupaciones.py)",
           f"Autogenerado desde el código · {len(ag.TODAS)} agrupaciones · "
           "portado de analisis_tasas/financiero/agrupaciones.py · "
           "los alias de esta hoja son los que usan las fórmulas de la hoja 2",
           HA, rows, [20, 38, 36, 12, 65, 60])

os.makedirs("docs", exist_ok=True)
out = os.path.join("docs", "diccionario_metricas.xlsx")
wb.save(out)
print(f"OK {out} · hojas: {wb.sheetnames} · agrupaciones: {len(rows)}")
